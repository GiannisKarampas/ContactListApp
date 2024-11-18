import openpyxl as op
import pandas as pd
import os
from openpyxl.styles import PatternFill

destination_path = r'testData/uploads/'
comparison_path = r'testData/comparison/'

class ExcelHandler:
    @staticmethod
    def read_excel_data(filename, sheet_name):
        # Load the excel file
        workbook = op.load_workbook(filename)

        # Select the worksheet
        worksheet = workbook[sheet_name]

        print(f'Total number of rows: {worksheet.max_row}')
        print(f'Total number of columns: {worksheet.max_column}')

        # Reading data from multiple cells
        headers = []
        for cell in worksheet[1]:
            headers.append(cell.value)
        
        values = [worksheet.cell(row=2, column=1).value for i in range(1, worksheet.max_column + 1)]
        print(headers)
        print(values)

        workbook.close()
    
    @staticmethod
    def update_cells(filename, sheet_name):
        updated_excel_file = os.path.join(destination_path, 'updated_excel_file.xlsx')
        # Load the excel file
        workbook = op.load_workbook(filename)

        # Select the worksheet
        worksheet = workbook[sheet_name]

        # Find the headers row index
        header_row = 1

        # Set new values for specific cells based on headers
        headers_to_change = ['header 1', 'header 2']
        new_values = [['new_value 1'], ['new_value 2']]
        headers_cells = [cell for cell in worksheet[header_row] if cell.value in headers_to_change]
        column_indexes = [cell.column for cell in headers_cells]

        for header_index, col in enumerate(column_indexes):
            for row in range(header_row + 1, worksheet.max_row + 1):
                worksheet.cell(row, col).value = new_values[header_index][row - header_row - 1]

        # Save the workbook
        workbook.save(updated_excel_file)
        print('Excel file has been created')
        workbook.close()

    @staticmethod
    def compare_excel_files(first_file, second_file):
        diff = False
        if os.path.splitext(first_file)[1] != '.xlsx' and os.path.splitext(second_file)[1] != '.xlsx':
            pd.read_csv(first_file).to_excel(f'{os.path.splitext(first_file)[0]}.xlsx', sheet_name='Sheet1', index=False)
            pd.read_csv(second_file).to_excel(f'{os.path.splitext(second_file)[0]}.xlsx', sheet_name='Sheet1', index=False)
            first_file = f'{os.path.splitext(first_file)[0]}.xlsx'
            second_file = f'{os.path.splitext(second_file)[0]}.xlsx'
        
        data1 = op.load_workbook(first_file)
        data2 = op.load_workbook(second_file)
        
        data1_sheet = data1.active
        data2_sheet = data2.active

        fill_format = PatternFill(start_color='FF5133', end_color='FF5133', fill_type='solid')

        for column in data1_sheet.iter_cols():
            for cell in column:
                current_cell_value = cell.value
                cell_location = cell.coordinate
                if current_cell_value != data2_sheet[cell_location].value:
                    cell.fill = fill_format
                    diff = True
        
        if diff:
            data1.save(os.path.join(comparison_path, 'comparison_file.xlsx'))
            print('The files have differences')
            print(f'The comparison file comparison_file.xlsx has been creted in the {comparison_path}')
            data1.close()